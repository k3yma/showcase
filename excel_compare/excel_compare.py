import pandas as pd
import numpy as np
import argparse
from openpyxl import formatting, styles
from datetime import datetime


def load_file(f):
    if f.endswith(".xls") or f.endswith(".xlsx"):
        func = pd.read_excel
    elif f.endswith(".csv"):
        func = pd.read_csv
    else:
        print(f"Invalid file type {f}")
        exit(1)

    df = func(f)

    if sort_key:
        df.sort_values(by=sort_key, inplace=True)

    return df.sort_index(axis=1).reset_index(drop=True).fillna("")


def compare_dfs_summary(df1, df2):
    return pd.DataFrame(
        df1.rename(columns={c: f"col{i}" for i, c in enumerate(df1.columns)})
        == df2.rename(columns={c: f"col{i}" for i, c in enumerate(df2.columns)})
    ).rename(columns={f"col{i}": c for i, c in enumerate(df1.columns)})


def compare_dfs_detail(df1, df2):
    cols = {f"col{i}": c for i, c in enumerate(df1.columns)}
    cols["self"] = "File 1"
    cols["other"] = "File 2"

    return (
        df1.rename(columns={c: f"col{i}" for i, c in enumerate(df1.columns)})
        .compare(df2.rename(columns={c: f"col{i}" for i, c in enumerate(df2.columns)}))
        .rename(columns=cols)
    )


if __name__ == "__main__":
    default_output_file = f"comparison_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    parser = argparse.ArgumentParser("Data Sheet Comparator")
    parser.add_argument("file1", help="File1")
    parser.add_argument("file2", help="File2")
    parser.add_argument("-s", dest="sort_key", help="Column name to sort files by")
    parser.add_argument("-o", dest="output_file", help="Output filename")
    args = parser.parse_args()

    file1 = args.file1
    file2 = args.file2
    sort_key = args.sort_key
    output_file = args.output_file or default_output_file

    assert output_file.endswith(".xlsx"), "Output file must be of type .xlsx"

    df1 = load_file(file1)
    df2 = load_file(file2)
    dfc = compare_dfs_summary(df1, df2)
    dfdd = compare_dfs_detail(df1, df2)
    dfd = dfc.replace(True, np.nan).dropna(how="all", axis=0).dropna(how="all", axis=1)

    # Write out the resulting file.
    writer = pd.ExcelWriter(output_file)

    df1.to_excel(writer, sheet_name="File 1")
    df2.to_excel(writer, sheet_name="File 2")
    dfc.to_excel(writer, sheet_name="Comparison")
    dfd.to_excel(writer, sheet_name="Differences")
    dfdd.to_excel(writer, sheet_name="Difference Detail")

    # Overview
    overview = writer.book.create_sheet("Overview", 0)
    overview["A1"] = "File 1"
    overview["B1"] = file1
    overview["A2"] = "File 2"
    overview["B2"] = file2
    overview["A3"] = "Headings Match"
    overview["B3"] = "Yes" if set(df1.columns) == set(df2.columns) else "No"

    # Add conditional formatting
    writer.sheets["Comparison"].conditional_formatting.add(
        "B2:XFD1048576",
        formatting.rule.FormulaRule(formula=["ISBLANK(B2)"], stopIfTrue=True),
    )

    writer.sheets["Comparison"].conditional_formatting.add(
        "B2:XFD1048576",
        formatting.rule.CellIsRule(
            operator="equal",
            formula=["True"],
            fill=styles.PatternFill(bgColor="c6efce"),
        ),
    )

    writer.sheets["Comparison"].conditional_formatting.add(
        "B2:XFD1048576",
        formatting.rule.CellIsRule(
            operator="equal",
            formula=["False"],
            fill=styles.PatternFill(bgColor="ffc7ce"),
        ),
    )

    writer.sheets["Differences"].conditional_formatting.add(
        "B2:XFD1048576",
        formatting.rule.FormulaRule(formula=["ISBLANK(B2)"], stopIfTrue=True),
    )

    writer.sheets["Differences"].conditional_formatting.add(
        "B2:XFD1048576",
        formatting.rule.CellIsRule(
            operator="equal",
            formula=["False"],
            fill=styles.PatternFill(bgColor="ffc7ce"),
        ),
    )

    # Freeze panes on all
    for w in writer.sheets:
        writer.sheets[w].freeze_panes = "B3" if w == "Difference Detail" else "B2"

    writer.save()

    print(f"{output_file} written.")
